from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import WorkbookAlreadySaved
from openpyxl.utils.exceptions import ReadOnlyWorkbookException
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors

shipping_costs = []
customer_row_nums = []
highlight_green = PatternFill(patternType='solid', fgColor=colors.Color(rgb='00FF55'))
highlight_blue = PatternFill(patternType='solid', fgColor=colors.Color(rgb='0dacfa'))

# Get order shipping costs and track unique orders.
def orderInfo(ws):
    for row in range(2, ws.max_row):
        for col in range(10, 11):
            char = get_column_letter(col)
            if ws[char + str(row)].value != None:
                shipping_costs.append(ws[char + str(row)].value)
                customer_row_nums.append(row)


# Format date to mm/dd/yyyy
def formatDate(ws):
    for row in customer_row_nums:
        try:
            # Enter date in cell if empty
            if ws['D' + str(row)].value == None:
                for col in range(16, 17):
                    char = get_column_letter(col)
                    cell_date = (ws[char + str(row)].value[0:10:1])
                    year, month, day = cell_date.split('-')
                    cell_date = month + "/" + day + "/" + year
                    ws['D' + str(row)].value = cell_date
                    ws['D' + str(row)].fill = highlight_blue
            else:
                for col in range(4, 5):
                    char = get_column_letter(col)
                    if ws[char + str(row)].value != None:
                        cell_date = (ws[char + str(row)].value[0:10:1])
                        year, month, day = cell_date.split('-')
                        cell_date = month + "/" + day + "/" + year
                        ws[char + str(row)].value = None
                        ws[char + str(row)].value = cell_date
                        ws[char + str(row)].fill = highlight_blue

        except TypeError:
            pass


'''
Insert new row before each unique order (except first order in spreadsheet)
Append new row to last order
Copy order number to new rows
Add shipping as a new "LineItem" and shipping cost
'''
def insertRows(ws):
    difference = 0
    try:
        for i in range(1, len(customer_row_nums)):
            ws.insert_rows(customer_row_nums[i] + difference)
            ws['A' + str(customer_row_nums[i] + difference)].value = ws['A' + str(customer_row_nums[i] + difference - 1)].value
            ws['S' + str(customer_row_nums[i] + difference)].value = shipping_costs[i - 1]
            ws['R' + str(customer_row_nums[i] + difference)].value = 'Shipping'
            ws['W' + str(customer_row_nums[i] + difference)].value = 'TRUE'
            customer_row_nums[i] = customer_row_nums[i] + difference
            difference = i

        customer_row_nums.pop(0)

        ws.append({'A': ws['A' + str(ws.max_row)].value,
                'S': shipping_costs[-1],
                'R': 'Shipping:SS',
                'W': 'TRUE'})

        customer_row_nums.append(ws.max_row)

        # Highlight added rows
        for row in customer_row_nums:
            for col in range(1, (ws.max_column + 1)):
                char = get_column_letter(col)
                ws[char + str(row)].fill = highlight_green

    except ReadOnlyWorkbookException as e:
        print("Error: Trying to modify a read-only workbook!")


def main():

    print("*** Enter file or type 'quit' to exit program ***")
    excel_file = input("Provide excel file or filepath (ex. test.xlsx): ")

    while excel_file != 'quit':
        try:
            print("\n-> Loading Excel File...")
            wb = load_workbook(excel_file)
            ws = wb.active
            orderInfo(ws)
            formatDate(ws)
            insertRows(ws)
            print("-> Modifying File...")

            try:
                wb.save("MOD_" + excel_file)
                print("*** Success! ***\n")

            except WorkbookAlreadySaved as e:
                print("Workbook Already Saved\n")

        except:
            print("\nERROR: Failed to load Excel File! Make sure the filename/filepath is correct.")
            print("Please re-run the script\n")

        print("*** Enter file or type 'quit' to quit ***")
        excel_file = input("Provide excel file or filepath (ex. test.xlsx): ")



if __name__ == '__main__':
    main()
